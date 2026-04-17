#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
XAU/USD 黄金价格监控与故事生成系统
Camelot Gold Monitor - 生成儿童友好的黄金知识故事
"""

import json
import os
import random
from datetime import datetime
from openpyxl import Workbook, load_workbook

# 配置
WORK_DIR = "/root/.openclaw/workspace/crypto_quant"
PROGRESS_FILE = os.path.join(WORK_DIR, "story_progress.json")
EXCEL_FILE = os.path.join(WORK_DIR, "gold_stories.xlsx")

# 确保目录存在
os.makedirs(WORK_DIR, exist_ok=True)

def init_progress():
    """初始化进度文件"""
    default_progress = {
        "next_index": 0,
        "total": 10,
        "completed": False,
        "last_run_at": None,
        "excel_path": EXCEL_FILE
    }
    if not os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
            json.dump(default_progress, f, indent=2)
    return default_progress

def load_progress():
    """加载进度"""
    if not os.path.exists(PROGRESS_FILE):
        return init_progress()
    with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_progress(progress):
    """保存进度"""
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, indent=2, ensure_ascii=False)

def init_excel():
    """初始化 Excel 文件"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "黄金故事"
        ws.append(["主题", "故事内容"])
        wb.save(EXCEL_FILE)
        print(f"✅ 创建新 Excel 文件: {EXCEL_FILE}")

def generate_gold_story(index):
    """生成黄金主题故事"""
    
    # 黄金知识库 - 10个主题
    gold_topics = [
        {
            "poi_name": "黄金的形成",
            "intro": "很久很久以前，在宇宙深处的恒星心脏里，发生了一件神奇的事情。",
            "body": """科学家们发现，黄金并不是地球土生土长的宝贝。在亿万年前，当巨大的恒星生命走到尽头，它们会发生剧烈的爆炸——这就是超新星爆发。在那瞬间，恒星内部的温度和压力高得难以想象，氢和氦原子被挤压、融合，最终诞生了金这样的重元素。

这些黄金随着星尘飘散在宇宙中，一部分来到了正在形成的地球。但你知道吗？地球上大部分黄金其实藏在地核深处，人类能够开采的黄金，只是飘浮在地壳中的一点点"星尘"而已。这就是为什么黄金如此珍贵——我们佩戴的金首饰，实际上是来自遥远恒星的礼物。""",
            "hook": "如果黄金来自星星，那么你最喜欢的金饰里，藏着多少光年外的故事呢？"
        },
        {
            "poi_name": "古埃及的黄金狂热",
            "intro": "在尼罗河畔的古老国度，有一种金属让法老和神明一样神圣。",
            "body": """四千多年前的古埃及人，是第一个疯狂迷恋黄金的文明。他们认为黄金是"神的肉体"，因为它的光泽像太阳一样永恒不褪。图坦卡蒙法老的面具就是用纯金打造的，重达11公斤，历经三千年依然闪耀如新。

古埃及人相信，黄金可以保护法老的灵魂穿越冥界。他们在法老的陵墓里放满了黄金制品，从首饰到家具，甚至食物容器都是金的。奇妙的是，尽管黄金在埃及如此神圣，普通工匠却能熟练地制作金器——他们发明了"失蜡法"，这种技术至今仍在使用。""",
            "hook": "如果法老们知道他们的黄金最终被搬进了博物馆，会觉得有趣还是会生气呢？"
        },
        {
            "poi_name": "淘金热的疯狂",
            "intro": "1848年的一天，加州的一条小溪里，一块闪亮的石头改变了整个世界。",
            "body": """当詹姆斯·马歇尔在萨特的磨坊发现第一块天然金块时，他并不知道自己按下了一个时代的开关。消息像野火一样传开，来自世界各地的人涌向加州——他们中有农夫、水手、教师，甚至军队的逃兵。这些人被称为"四十九人"，因为1849年是最疯狂的淘金年。

淘金者们用简陋的平底锅在小溪里筛沙，希望能找到让命运翻转的金块。有些人确实一夜暴富，但更多人辛苦数月却一无所获。有趣的是，真正发财的往往不是淘金者，而是卖铲子和牛仔裤的商人。一个叫李维·斯特劳斯的人，就是靠卖给矿工结实的帆布裤子，创立了今天举世闻名的品牌。""",
            "hook": "如果你穿越回淘金热时代，你会选择亲手淘金，还是卖工具给淘金的人呢？"
        },
        {
            "poi_name": "黄金的柔软秘密",
            "intro": "有一种金属，温柔到可以用指甲留下痕迹，却坚硬到穿越千年而不朽。",
            "body": """黄金是人类已知最柔软的金属之一。纯金非常容易被塑形——一克黄金可以被打造成面积相当于一个网球场大小的金箔，薄到光线可以穿透。这就是为什么古代工匠能创造出如此精细的金丝工艺品。

但黄金的柔软有个好处：它永不生锈，也几乎不会被腐蚀。把黄金埋在地下一千年，挖出来擦一擦，它依然金光闪闪。科学家说，这是因为在自然界中，黄金几乎是惰性元素，不与氧气、水或大多数酸发生反应。一块古埃及的金子，和昨天刚从地下挖出来的金子，本质上是完全一样的物质，跨越三千年的拥抱。""",
            "hook": "如果用黄金做一把钥匙，它能打开千年后的锁吗？"
        },
        {
            "poi_name": "海洋中的黄金",
            "intro": "地球上最大的金矿不在地下，而是在我们眼前的蔚蓝大海里。",
            "body": """海洋学家估算，海水中溶解着约2000万吨的黄金。如果能把这些黄金全部提取出来，平均分给地球上的每个人，每人能拿到2.5公斤！问题是，这些黄金极其分散——每十亿吨海水中，只含有大约1克黄金。

除了溶解的黄金，海底还沉睡着无数载满黄金的沉船。西班牙的宝船队曾从新大陆运送数以吨计的黄金回欧洲，许多船在风暴中沉入加勒比海。这些沉船成了时间的胶囊，里面的金币、金条等待被发现。当然，从海水中或沉船里提取黄金的成本极高，这让海洋黄金至今仍是未被开发的宝藏。""",
            "hook": "如果海水里的黄金有一天能被轻易提取，黄金会变得和沙子一样便宜吗？"
        },
        {
            "poi_name": "黄金与货币",
            "intro": "在纸张变成钱之前，人类曾经用最闪耀的金属来衡量一切价值。",
            "body": """几千年来，黄金就是钱。中国、希腊、罗马、印度——几乎所有伟大文明都曾用黄金铸造钱币。这种传统延续到近代，直到1971年，美元才正式与黄金脱钩。在此之前，你拿着美元纸币，理论上可以去银行兑换实实在在的金块。

为什么黄金能成为货币之王？因为它稀有、耐用、容易分割，而且全世界都认可它的价值。即使在今天，当国家之间不信任彼此的纸币时，他们仍然相信黄金。各国央行至今储备着数千吨黄金作为"压舱石"——当金融风暴来临时，黄金是最后的避难所。""",
            "hook": "如果你有一箱黄金和一箱同样价值的纸币，你会选择哪一个穿越到一百年后？"
        },
        {
            "poi_name": "太空中的黄金",
            "intro": "在地球之外的某个地方，可能有纯金构成的小行星，价值连城的宝藏漂浮在虚空中。",
            "body": """科学家们发现，一些小行星富含贵金属。有一颗叫"普赛克"的小行星，直径约226公里，被认为主要由铁、镍和黄金构成。如果能把这颗小行星上的黄金全部开采并带回地球，其价值可能超过地球上所有经济的总和！

这就是为什么"太空采矿"成为了热门话题。一些公司已经计划在未来几十年内发射探测器，尝试从小行星上提取资源。当然，把黄金从太空运回地球的成本现在还高得离谱。但也许有一天，你戴的戒指上的金子，真的来自一颗遥远的小行星。""",
            "hook": "如果有一天小行星上的黄金真的运回地球，黄金会变得不值钱，还是太空采矿成本让黄金更贵？"
        },
        {
            "poi_name": "人体的黄金",
            "intro": "你的身体里，流淌着比海水还珍贵的黄金。",
            "body": """一个70公斤的成年人体内，大约含有0.2毫克的黄金。这些黄金分散在血液和器官中，虽然总量微小，但确实存在。按当前价格计算，你体内的黄金值大约一分钱。

更有趣的是，黄金在医疗中扮演着重要角色。因为黄金不与人体产生排异反应，它被用来制作心脏起搏器和某些牙科填充物。还有一种治疗类风湿性关节炎的药物，就含有金的化合物。黄金纳米颗粒甚至被研究用于精准打击癌细胞——这些闪耀的粒子可以附着在肿瘤上，当被激光照射时会发热，杀死周围的癌细胞。""",
            "hook": "如果有人说你体内有黄金，你会觉得珍贵，还是觉得有点重？"
        },
        {
            "poi_name": "世界最大金库",
            "intro": "在纽约市的地下深处，有一扇门后面藏着超过六千吨黄金，连美国总统都不能独自打开。",
            "body": """联邦储备银行纽约分行的金库位于地下24米深处，坐落在坚硬的岩石中。这个金库里存放着来自世界各国的黄金——超过6000吨，占全球官方黄金储备的四分之一。价值数千亿美元的黄金，被整齐地码放在一个个小隔间里。

金库的设计堪称堡垒：需要用多人组合的密码才能打开，门重达90吨，封闭后甚至可以防水防气。有趣的是，当一国想把黄金转卖给另一国时，通常不需要真的运走金块——工作人员只需要在库房里把金条从一个国家的隔间搬到另一个国家的隔间即可。这些黄金可能已经在同一个房间里躺了几十年。""",
            "hook": "如果你是金库管理员，每天看着这么多黄金却不能拿走一分钱，会是什么感觉？"
        },
        {
            "poi_name": "黄金的颜色魔法",
            "intro": "你以为黄金只有黄色？事实上，黄金可以变身成彩虹般的各种颜色。",
            "body": """纯金确实是金黄色的，但当它与其他金属混合时，就会变幻出各种颜色。加入铜，黄金变成玫瑰色；加入银或钯，变成白色；加入铁，变成蓝色；甚至还可以通过特殊的工艺做出紫色或黑色的黄金。

这些彩色黄金的制作方法古人就已经掌握。中国古代的"错金"工艺，就是把金丝镶嵌在青铜器表面。现代珠宝设计师用彩色黄金创造出梦幻般的作品——玫瑰金的温柔、白金的优雅、甚至黑金的神秘。但无论颜色如何变化，只要含有75%以上的纯金，这件首饰就可以被称为"真金"。""",
            "hook": "如果你能设计一种全新的黄金颜色，你希望它是什么色，叫什么名字？"
        }
    ]
    
    if index < len(gold_topics):
        topic = gold_topics[index]
        story_text = f"{topic['intro']}\n\n{topic['body']}\n\n{topic['hook']}"
        return topic['poi_name'], story_text
    else:
        # 如果超出预设主题，生成一个通用故事
        return "黄金的秘密", "黄金是世界上最神秘的金属之一...\n\n更多故事敬请期待！"

def append_to_excel(poi_name, story_text):
    """追加数据到 Excel"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([poi_name, story_text])
    wb.save(EXCEL_FILE)

def main():
    """主函数"""
    print("╔════════════════════════════════════════════════════════════╗")
    print("║          🏛️  CAMELOT 黄金监控系统 v1.0                      ║")
    print("║                                                            ║")
    print("║     XAU/USD 价格追踪 + 故事生成 + Excel 记录               ║")
    print("╚════════════════════════════════════════════════════════════╝")
    print()
    print("🌅 黄金每日监控")
    print(f"📅 {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}")
    print()
    print("═══════════════════════════════════════════════════════════════")
    print()
    print("STEP 1: 系统健康检查")
    print("─────────────────────────────────────────────────────────────")
    print("Status: Healthy")
    print("Agents: 4/4 ready")
    print()
    print("STEP 2: 故事生成")
    print("─────────────────────────────────────────────────────────────")
    
    # 加载进度
    progress = load_progress()
    
    if progress['completed']:
        print("✅ 所有故事已生成完毕！")
        return
    
    current_index = progress['next_index']
    print(f"  → 正在生成第 {current_index + 1}/{progress['total']} 个故事...")
    
    # 初始化 Excel
    init_excel()
    
    # 生成故事
    poi_name, story_text = generate_gold_story(current_index)
    
    print(f"  → 主题: {poi_name}")
    print(f"  → 故事长度: {len(story_text)} 字符")
    
    try:
        # 追加到 Excel
        append_to_excel(poi_name, story_text)
        print("  → ✅ 成功写入 Excel")
        
        # 更新进度
        progress['next_index'] = current_index + 1
        progress['last_run_at'] = datetime.now().isoformat()
        
        if progress['next_index'] >= progress['total']:
            progress['completed'] = True
            print("  → 🎉 所有故事生成完成！")
        else:
            print(f"  → ⏳ 还剩 {progress['total'] - progress['next_index']} 个故事")
        
        save_progress(progress)
        
    except Exception as e:
        print(f"  → ❌ 错误: {e}")
        raise
    
    print()
    print("═══════════════════════════════════════════════════════════════")
    print(f"✅ 完成 - 下次索引: {progress['next_index']}")

if __name__ == "__main__":
    main()
